import asyncio
import time
from datetime import datetime, timedelta
import aioschedule
from aiogram import Bot, executor, types
from aiogram.dispatcher import Dispatcher
from aiogram.dispatcher.webhook import SendMessage
from aiogram.utils.executor import start_webhook
from openpyxl import Workbook, load_workbook
from telegram import InlineKeyboardMarkup, TelegramObject
import os
from flask import Flask, request

TOKEN = ''
APP_URL = f"https://cleanhometelebot.herokuapp.com/{TOKEN}"

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
server = Flask(__name__)

@server.route('/' + TOKEN, methods=['POST'])
def get_message():
    json_string = request.get_data().decode('utf-8')
    update = types.Update.de_json(json_string)
    bot.process_new_updates([update])
    return '!',200

@server.route('/')
def webhook():
    bot.delete_webhook()
    bot.set_webhook(url=APP_URL)
    return '!',200



    
@dp.message_handler()
async def command_handler(message: types.Message()):
    cleaning_log_wb = load_workbook('cleaning_log.xlsx')
    clsx = cleaning_log_wb['Sheet1']
    if message.text == '/points@PariglaHomebot':
        points_message = "User1: {} \nUser2: {}".format(clsx.cell(row= 2,column= 10).value,clsx.cell(row= 3,column= 10).value)
        await bot.send_message(chat_id=message.chat.id,text=points_message)

@dp.callback_query_handler()
async def answer(call):
    print(call)
    cleaning_log_wb = load_workbook('cleaning_log.xlsx')
    clsx = cleaning_log_wb['Sheet1']
    name = call.from_user.first_name
    number_of_activities = 0
    for cell in clsx['A']:
        if cell.value != None:
            number_of_activities += 1
    
    for activity in range(2,number_of_activities+1):
        if clsx.cell(row= activity,column= 3).value == call.data:
            clsx.cell(row= activity,column= 1, value= datetime.now())

            if name == 'sun.dealer':
               current_points =  clsx.cell(row= 2,column= 10).value
               activity_points = clsx.cell(row= activity,column= 8).value
               clsx.cell(row= 2,column= 10, value= activity_points+current_points)
            if name == 'Alexandra':
               current_points =  clsx.cell(row= 3,column= 10).value
               activity_points = clsx.cell(row= activity,column= 8).value
               clsx.cell(row= 3,column= 10, value= activity_points+current_points)
            if call.data == 'clothes_washing':
                clsx.cell(row= activity,column= 2, value= 'Вытащить одежду')
                clsx.cell(row= activity,column= 3, value= "clothes_out")
                clsx.cell(row= activity,column= 4, value= 0) #недели
                clsx.cell(row= activity,column= 5, value= 0) #дни
                clsx.cell(row= activity,column= 6, value= 12) #часы
                clsx.cell(row= activity,column= 7, value= 0) #минуты
                clsx.cell(row= activity,column= 8, value= 1) #баллы
            if call.data == 'clothes_out':
                clsx.cell(row= activity,column= 2, value= 'Разобрать одежду')
                clsx.cell(row= activity,column= 3, value= "clothes_sorted")
                clsx.cell(row= activity,column= 4, value= 0) #недели
                clsx.cell(row= activity,column= 5, value= 0) #дни
                clsx.cell(row= activity,column= 6, value= 0) #часы
                clsx.cell(row= activity,column= 7, value= 10) #минуты
                clsx.cell(row= activity,column= 8, value= 1) #баллы
            if call.data == 'clothes_sorted':
                clsx.cell(row= activity,column= 2, value= 'Закинуть одежду стираться')
                clsx.cell(row= activity,column= 3, value= "clothes_washing")
                clsx.cell(row= activity,column= 4, value= 0) #недели
                clsx.cell(row= activity,column= 5, value= 0) #дни
                clsx.cell(row= activity,column= 6, value= 0) #часы
                clsx.cell(row= activity,column= 7, value= 10) #минуты
                clsx.cell(row= activity,column= 8, value= 1) #баллы
            if call.data == 'dishes_collected':
                clsx.cell(row= activity,column= 2, value= 'загрузить посуду')
                clsx.cell(row= activity,column= 3, value= "dishes_washing")
                clsx.cell(row= activity,column= 4, value= 0) #недели
                clsx.cell(row= activity,column= 5, value= 0) #дни
                clsx.cell(row= activity,column= 6, value= 0) #часы
                clsx.cell(row= activity,column= 7, value= 10) #минуты
                clsx.cell(row= activity,column= 8, value= 1) #баллы
            if call.data == 'dishes_washing':
                clsx.cell(row= activity,column= 2, value= 'разобрать посуду')
                clsx.cell(row= activity,column= 3, value= "dishes_sorted")
                clsx.cell(row= activity,column= 4, value= 0) #недели
                clsx.cell(row= activity,column= 5, value= 0) #дни
                clsx.cell(row= activity,column= 6, value= 5) #часы
                clsx.cell(row= activity,column= 7, value= 0) #минуты
                clsx.cell(row= activity,column= 8, value= 1) #баллы
            if call.data == 'dishes_sorted':
                clsx.cell(row= activity,column= 2, value= 'собрать посуду со всего дома')
                clsx.cell(row= activity,column= 3, value= "dishes_collected")
                clsx.cell(row= activity,column= 4, value= 0) #недели
                clsx.cell(row= activity,column= 5, value= 0) #дни
                clsx.cell(row= activity,column= 6, value= 0) #часы
                clsx.cell(row= activity,column= 7, value= 10) #минуты
                clsx.cell(row= activity,column= 8, value= 1) #баллы
            cleaning_log_wb.save('cleaning_log.xlsx')




    if len(call.message.reply_markup['inline_keyboard']) == 1:
        await bot.delete_message(chat_id= call.message.chat.id, message_id= call.message.message_id)
        await bot.send_message("-1001223803040", "Поздравляю, дела по дому  на сегодня окончены!!!")
    else:
        current_markup = call.message.reply_markup['inline_keyboard']
        new_inline_makrup = types.InlineKeyboardMarkup()
        for button in current_markup:
            if button[0].callback_data != call.data:
                new_inline_makrup.add(types.InlineKeyboardButton(text = button[0].text,
                                                                 callback_data = button[0].callback_data))
                await bot.edit_message_reply_markup(chat_id=call.message.chat.id,
                                                    message_id=call.message.message_id,
                                                    reply_markup = new_inline_makrup)



async def morning_tasks_send():
    markup_inline = types.InlineKeyboardMarkup()
    cleaning_log_wb = load_workbook('cleaning_log.xlsx')
    clsx = cleaning_log_wb['Sheet1']
    number_of_activities = 0
    for cell in clsx['A']:
        if cell.value != None:
            number_of_activities += 1
    for activity in range(2,number_of_activities+1):
        current_timedelta = timedelta(weeks= clsx.cell(row= activity, column= 4).value,
                                      days = clsx.cell(row= activity, column= 5).value,
                                      hours= clsx.cell(row= activity, column= 6).value,
                                      minutes= clsx.cell(row= activity, column= 7).value)
        if datetime.now() - clsx.cell(row= activity, column= 1).value > current_timedelta:
            markup_inline.add(types.InlineKeyboardButton(text = clsx.cell(row= activity, column= 2).value,callback_data = clsx.cell(row= activity, column= 3).value))

    await bot.send_message("-1001223803040", "Задачи на это утро",reply_markup = markup_inline)


async def evening_tasks_send():
    await bot.send_message("-1001223803040", "byewrld")


async def scheduler():
    aioschedule.every().day.at("5:28").do(morning_tasks_send)
    aioschedule.every().day.at("16:49").do(evening_tasks_send)
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(1)

async def on_startup(_):
    asyncio.create_task(scheduler())


if __name__ == '__main__':
    server.run(host='0.0.0.0',port=int(os.environ.get('PORT', 5000)))
    executor.start_polling(dp, skip_updates=False, on_startup=on_startup)





# schedule.every(10).minutes.do(job)
# schedule.every().hour.do(job)
# schedule.every().day.at("10:30").do(job)
# schedule.every(5).to(10).minutes.do(job)
# schedule.every().monday.do(job)
# schedule.every().wednesday.at("13:15").do(job)
# schedule.every().minute.at(":17").do(job)
